import dto.Product as prDto
import dto.User as userDTO
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.styles import Alignment
import random
import datetime
import base64
import requests
productDB = prDto.Product({'host': 'j45316134.myjino.ru', 'username': 'j45316134', 'password': 'cf}j}R75tRzM', 'dbname': 'j45316134', })
userDB = userDTO.User({'host': 'j45316134.myjino.ru', 'username': 'j45316134', 'password': 'cf}j}R75tRzM', 'dbname': 'j45316134', })
tgLink =  "https://api.telegram.org/bot5578668426:AAGN8on9LTVIayUOfPSerT0RjkqzGIKNC1A/sendMessage?chat_id={0}&text={1}"
answerMorning = '''👋 Добрый день! 
👋 Добрый день! 
 На сегодня у нас следующий план:

💳Будут заказаны следующие товары: 
{0}

🛍 Ожидают получения на пункте выдачи следующие товары: 
{1}

📦 Полученные товары:
{2}

📝 Отзывы ожидающие согласования:
{3}
                
⏰ Отзывы которые поставлены на публикацию:
{4}

⏰ Отзывы которые опубликованы:
{5}

✅ Отзывы которые успешно прошли модерацию
{6}

‼️ Удалённые и ошибки при публикации отзывов:
{7}

‼️Напоминаем вам, что по получению данного уведомления, изменить кол-во заказов на сегодня уже нельзя. Так как все изменения мы принимаем за 24 часа.‼️

С уважением компания RATE-THIS, хорошего дня!'''

answerNight = '''Доброй ночи!
Сверка выполненного плана:

💳 За сегодня было сделано заказов: 
{0}

📦 Полученных товаров
{1}

📝 Опубликованных отзывов 
{2}'''


class Product:
    def __init__(self):
        self.taskUser = {}
        pass

    def setTask(self, task, tg):
        self.taskUser[tg] = task
        print(self.taskUser[tg])
    def dayPush(self):
        answerList = []
        user = productDB.getUsersNot()
        for j in user:

            task = j[0]
            result = productDB.buyout(task, '0,1',  datetime.date.today().strftime("%y-%m-%d"))
            answerList.append('')
            for i in result:
                answerList[len(answerList) - 1] += "    {0}: {1}шт\n".format(i[1], i[0])
            result = productDB.buyout(task, '2,3', False)
            answerList.append('')
            for i in result:
                answerList[len(answerList) - 1] += "    {0}: {1}шт\n".format(i[1], i[0])
            result = productDB.buyout(task, '4,5,6,7,8', False)
            answerList.append('')
            for i in result:
                answerList[len(answerList) - 1] += "    {0}: {1}шт\n".format(i[1], i[0])
            result = productDB.review(task, '4', False)
            answerList.append('')
            for i in result:
                answerList[len(answerList) - 1] += "    {0}: {1}шт\n".format(i[1], i[0])
            result = productDB.review(task, '5', False)
            answerList.append('')
            for i in result:
                answerList[len(answerList) - 1] += "    {0}: {1}шт\n".format(i[1], i[0])
            result = productDB.review(task, '6', False)
            answerList.append('')
            for i in result:
                answerList[len(answerList) - 1] += "    {0}: {1}шт\n".format(i[1], i[0])
            result = productDB.review(task, '7', False)
            answerList.append('')
            for i in result:
                answerList[len(answerList) - 1] += "    {0}: {1}шт\n".format(i[1], i[0])
            result = productDB.review(task, '8', False)
            answerList.append('')
            for i in result:
                answerList[len(answerList) - 1] += "    {0}: {1}шт\n".format(i[1], i[0])
            if j[1]:
                print(j[1])
                requests.get(tgLink.format(j[1], answerMorning.format(answerList[0], answerList[1], answerList[2], answerList[3], answerList[4], answerList[5], answerList[6], answerList[7])))

    def dayPushOne(self, task, tg):
        answerList = []
        result = productDB.buyout(task, '0,1', False)
        answerList.append('')
        for i in result:
            answerList[len(answerList) - 1] += "    {0}: {1}шт\n".format(i[1], i[0])
        result = productDB.buyout(task, '2,3', False)
        answerList.append('')
        for i in result:
            answerList[len(answerList) - 1] += "{0}: {1}шт\n".format(i[1], i[0])
        result = productDB.buyout(task, '4,5,6,7,8', False)
        print(result)
        answerList.append('')
        for i in result:
            answerList[len(answerList) - 1] += "    {0}: {1}шт\n".format(i[1], i[0])
        result = productDB.review(task, '4', False)
        answerList.append('')
        for i in result:
            answerList[len(answerList) - 1] += "    {0}: {1}шт\n".format(i[1], i[0])
        result = productDB.review(task, '5', False)
        answerList.append('')
        for i in result:
            answerList[len(answerList) - 1] += "    {0}: {1}шт\n".format(i[1], i[0])
        result = productDB.review(task, '6', False)
        answerList.append('')
        for i in result:
            answerList[len(answerList) - 1] += "    {0}: {1}шт\n".format(i[1], i[0])
        result = productDB.review(task, '7', False)
        answerList.append('')
        for i in result:
            answerList[len(answerList) - 1] += "    {0}: {1}шт\n".format(i[1], i[0])
        result = productDB.review(task, '8', False)
        answerList.append('')
        for i in result:
            answerList[len(answerList) - 1] += "    {0}: {1}шт\n".format(i[1], i[0])


        if tg:
            print(tg)
            requests.get(tgLink.format(tg, answerMorning.format(answerList[0], answerList[1], answerList[2],
                                                                  answerList[3], answerList[4], answerList[5],
                                                                  answerList[6], answerList[7])))



    def nightPush(self):
        date = datetime.date.today().strftime("%y-%m-%d")
        answerList = []
        user = productDB.getUsersNot()
        print(user)
        for j in user:
            answerList = []
            task = j[0]
            if j[1] in self.taskUser:
                task = self.taskUser[j[1]]

            result = productDB.buyout(task, '2,3', date)
            print(result)
            answerList.append('')
            sumup = 0
            for i in result:
                answerList[len(answerList) - 1] += "{0}: {1}шт\n".format(i[1], i[0])
                sumup+=i[2]


            answerList[len(answerList) - 1] += "🔄Всего было реализовано:\n".format(sumup)
            result = productDB.delivery(task, '4,5,6,7', date)
            answerList.append('')
            sumup = 0
            for i in result:
                answerList[len(answerList) - 1] += "{0}: {1}шт\n".format(i[1], i[0])
            print(answerList)
            result = productDB.review(task, '6,7,8', date)
            answerList.append('')
            sumup = 0
            for i in result:
                answerList[len(answerList) - 1] += "{0}: {1}шт\n".format(i[1], i[0])
            if j[1]:
                print(answerList)
                requests.get(tgLink.format(j[1], answerNight.format(answerList[0], answerList[1], answerList[2])))

    def reports(self, tg, type, dateType, date = False):
        status = {
            None: "Запланированно",
            0: "Запланированно",
            "": "Запланированно",
            1: "Корзина собирается",
            2: "Оплачен",
            3: "На пвз",
            4: "Получен",
            5: "Получен",
            6: "На модерации",
            7: "Опубликован",
            8: "Удалён",
            9: "Запланированно",
            10: "Запланированно",
            11: "Запланированно",
        }
        listName = {
            1: "Выкупам",
            2: "Доставкам",
            3: "Отзывам",
            4: '''Общий отчет за все время, в данном отчете, вы можете посмотреть статус по каждому вашему товару, что с ним на каком он этапе. Тип выкуп означает, что мы сделаем только выкуп товара, а тип отзыв, означает что сделаем выкуп и оставим отзыв''',
        }
        listDate = {
            1: datetime.date.today().strftime("%y-%m-%d"),
            2: "c {0} по {1}".format((datetime.date.today() - datetime.timedelta(days=7)).strftime("%y-%m-%d"), datetime.date.today().strftime("%y-%m-%d")),
            3: "c {0} по {1}".format((datetime.date.today() - datetime.timedelta(days=30)).strftime("%y-%m-%d"), datetime.date.today().strftime("%y-%m-%d")),
            4: "за {0}".format(date),
            5: "За все время",
        }
        task1 = self.taskUser[tg]
        # sheet.cells_dimensions[5].height = 28
        answerT = productDB.report(type, task1, dateType, date)
        answer = []
        if dateType > 1 and type < 3 and dateType < 4:

            for i in answerT:
                answer.append(productDB.getSortProduct(type, i[0], dateType, task1))
        else:
            answer = productDB.report(type, task1, dateType, date)

        workbook = Workbook()
        sheet = workbook.active
        sheet['A1'].alignment = Alignment(wrap_text=True)
        sheet.merge_cells('A1:C1')
        sheet['A1'] = 'RATE THIS\nPROMOTION'
        sheet['A1'].font = Font(name='Montserrat', size=24, color='46bdc6', b=True)
        sheet.column_dimensions['C'].width = 0000
        sheet.row_dimensions[1].height = 65
        sheet.merge_cells('D1:E1')

        sheet['D1'].alignment = Alignment(wrap_text=True)

        sheet[
            'D1'] = 'Лучший сервис по работе с маркетплейсами!\nТелефон для связи +7 (499) 133-39-37\nTelegram : {}\nСайт : https://rate-this.ru/'.format(
            "https://t.me/RATE_THISbot")
        sheet['D1'].font = Font(name='Montserrat', size=9, b=True)
        sheet.column_dimensions['E'].width = 28

        sheet.merge_cells('A2:E4')
        sheet['A2'] = 'Отчет по {0} {1}'.format(listName[type], listDate[dateType])
        sheet['A2'].font = Font(name='Montserrat', size=13, b=True)
        sheet['A2'].alignment = Alignment(wrap_text=True)
        if type < 3:
            sheet['A5'] = 'Дата'
            sheet['A5'].font = Font(name='Montserrat', size=9, b=True)

            sheet.column_dimensions['A'].width = 13

            sheet.column_dimensions['B'].width = 28

            sheet['B5'] = 'Артикул'
            sheet['B5'].font = Font(name='Montserrat', size=9, b=True)

            sheet['C5'] = 'Количество'
            sheet['C5'].font = Font(name='Montserrat', size=9, b=True)

            sheet.column_dimensions['D'].width = 28

            sheet['D5'] = 'Наименование'
            sheet['D5'].font = Font(name='Montserrat', size=9, b=True)

            sheet['E5'] = 'Стоимость'
            sheet['E5'].font = Font(name='Montserrat', size=9, b=True)
            sheet['E5'].alignment = Alignment(wrap_text=True)

            sheet['F5'] = 'Сумма'
            sheet['F5'].font = Font(name='Montserrat', size=9, b=True)
            sheet['F5'].alignment = Alignment(wrap_text=True)
        elif type < 4:
            sheet['A5'] = 'Дата'
            sheet['A5'].font = Font(name='Montserrat', size=9, b=True)

            sheet.column_dimensions['A'].width = 13

            sheet.column_dimensions['B'].width = 28

            sheet['B5'] = 'Статус'
            sheet['B5'].font = Font(name='Montserrat', size=9, b=True)

            sheet['C5'] = 'Артикул'
            sheet['C5'].font = Font(name='Montserrat', size=9, b=True)

            sheet.column_dimensions['D'].width = 28

            sheet['D5'] = 'Текст отзыва'
            sheet['D5'].font = Font(name='Montserrat', size=9, b=True)

            sheet['E5'] = 'Наименование'
            sheet['E5'].font = Font(name='Montserrat', size=9, b=True)
            sheet['E5'].alignment = Alignment(wrap_text=True)

            sheet['F5'] = 'Скрин отзыва'
            sheet['F5'].font = Font(name='Montserrat', size=9, b=True)
            sheet['F5'].alignment = Alignment(wrap_text=True)
        # sheet.cells_dimensions[5].height = 28
        else:
            sheet['A5'] = 'Статус'
            sheet['A5'].font = Font(name='Montserrat', size=9, b=True)

            sheet.column_dimensions['A'].width = 13

            sheet.column_dimensions['B'].width = 28

            sheet['B5'] = 'График выкупов'
            sheet['B5'].font = Font(name='Montserrat', size=9, b=True)

            sheet['C5'] = 'тип'
            sheet['C5'].font = Font(name='Montserrat', size=9, b=True)

            sheet.column_dimensions['D'].width = 28

            sheet['D5'] = 'Артикул'
            sheet['D5'].font = Font(name='Montserrat', size=9, b=True)

            sheet['E5'] = 'Размер'
            sheet['E5'].font = Font(name='Montserrat', size=9, b=True)
            sheet['E5'].alignment = Alignment(wrap_text=True)

            sheet['F5'] = 'Ключевой запрос'
            sheet['F5'].font = Font(name='Montserrat', size=9, b=True)
            sheet['F5'].alignment = Alignment(wrap_text=True)

            sheet['G5'] = 'Бренд'
            sheet['G5'].font = Font(name='Montserrat', size=9, b=True)
            sheet['G5'].alignment = Alignment(wrap_text=True)

            sheet['H5'] = 'Наименование'
            sheet['H5'].font = Font(name='Montserrat', size=9, b=True)
            sheet['H5'].alignment = Alignment(wrap_text=True)

            sheet['I5'] = 'Дата покупки'
            sheet['I5'].font = Font(name='Montserrat', size=9, b=True)
            sheet['I5'].alignment = Alignment(wrap_text=True)

            sheet['J5'] = 'Дата получения'
            sheet['J5'].font = Font(name='Montserrat', size=9, b=True)
            sheet['J5'].alignment = Alignment(wrap_text=True)

            sheet['K5'] = 'цена'
            sheet['K5'].font = Font(name='Montserrat', size=9, b=True)
            sheet['K5'].alignment = Alignment(wrap_text=True)

            sheet['L5'] = 'текст отзыва'
            sheet['L5'].font = Font(name='Montserrat', size=9, b=True)
            sheet['L5'].alignment = Alignment(wrap_text=True)

            sheet['M5'] = 'Скрин отзыва'
            sheet['M5'].font = Font(name='Montserrat', size=9, b=True)
            sheet['M5'].alignment = Alignment(wrap_text=True)

            sheet['N5'] = 'Дата отзыва'
            sheet['N5'].font = Font(name='Montserrat', size=9, b=True)
            sheet['N5'].alignment = Alignment(wrap_text=True)
        sheet.row_dimensions[5].height = 40
        index = 6
        sumBuy = 0
        sumPrice = 0
        print(answer)
        for i in answer:
            if dateType > 1 and type < 3 and dateType < 4:
                for j in i:
                    sheet[f'A{index}'] = str(j[1]).split(' ')[0]

                    sheet[f'B{index}'] = j[0]

                    sheet[f'C{index}'] = j[2]

                    sheet[f'D{index}'] = j[3]

                    sheet[f'E{index}'] = j[4]
                    try:
                        sheet[f'F{index}'] = int(j[4]) * int(j[2])
                    except Exception as e:
                        sheet[f'F{index}'] = "Не определенная цена"

                    try:
                        sumBuy += int(j[2])
                        sumPrice += int(j[4]) * int(j[2])
                    except Exception as e:
                        sumBuy = "не определенно"
                        sumPrice = "не определенно"

                    index += 1
            else:
                if type < 3:
                    sheet[f'A{index}'] = str(i[1]).split(' ')[0]

                    sheet[f'B{index}'] = i[0]

                    sheet[f'C{index}'] = i[2]

                    sheet[f'D{index}'] = i[3]

                    sheet[f'E{index}'] = i[4]
                    try:
                        sheet[f'F{index}'] = int(i[4]) * int(i[2])
                    except Exception as e:
                        sheet[f'F{index}'] = "Не определенная цена"

                    try:
                        sumBuy += int(i[2])
                        sumPrice += int(i[4]) * int(i[2])
                    except Exception as e:
                        sumBuy = "не определенно"
                        sumPrice = "не определенно"
                elif type < 4:
                    sheet[f'A{index}'] = str(i[0]).split(' ')[0]

                    sheet[f'B{index}'] = status[i[1]]

                    sheet[f'C{index}'] = i[2]

                    sheet[f'D{index}'] = base64.b64decode(i[3])

                    sheet[f'E{index}'] = i[4]

                    sheet[f'F{index}'] = i[5]
                else:
                    sheet[f'A{index}'] = status[i[0]]

                    sheet[f'B{index}'] = i[1]

                    sheet[f'C{index}'] = i[2]

                    sheet[f'D{index}'] = i[3]

                    sheet[f'E{index}'] = i[4]

                    sheet[f'F{index}'] = i[5]

                    sheet[f'G{index}'] = i[6]

                    sheet[f'H{index}'] = i[7]

                    sheet[f'I{index}'] = str(i[8]).replace(" ", "")

                    sheet[f'J{index}'] = str(i[9]).replace(" ", "")

                    sheet[f'K{index}'] = i[10]

                    sheet[f'L{index}'] = i[11]

                    sheet[f'M{index}'] = i[12]



                index+=1
        if type < 3:
            sheet[f'B{index + 1}'] = 'Итого'
            sheet[f'B{index + 1}'].font = Font(b=True)

            sheet[f'C{index + 1}'] = F"{sumBuy} шт."
            sheet[f'C{index + 1}'].font = Font(b=True)

            sheet[f'D{index + 1}'] = 'Сумма заказов'
            sheet[f'D{index + 1}'].font = Font(b=True)

            sheet[f'E{index + 1}'] = f'{sumPrice}'
            sheet[f'E{index + 1}'].font = Font(b=True)
        path = './Отчет по {1} save_{0}.xlsx'.format(random.randint(1111111111111, 9999999999999), listName[type])
        workbook.save(path)
        return path